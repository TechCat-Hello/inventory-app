from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import InventoryItem, Rental, ReturnLog
from django.http import HttpResponseForbidden
from .forms import InventoryItemForm, ItemSearchForm, RentalForm
from django.core.exceptions import PermissionDenied
from django.views.generic.edit import DeleteView
from django.urls import reverse_lazy
from django.db.models import Q
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth.models import User
import csv
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from django.template.loader import render_to_string
from weasyprint import HTML, CSS
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models.functions import TruncMonth
from django.db.models import Count, Sum
from django.utils.timezone import now, localtime
import calendar
from collections import defaultdict
from django.views import View
from datetime import datetime, timedelta
from django.conf import settings
import os
from weasyprint.text.fonts import FontConfiguration   #確認用
import logging    #確認用


# 一般ユーザー判定関数
def is_general_user(user):
    return user.is_authenticated and not user.is_staff

# 管理者判定関数
def is_admin(user):
    return user.is_superuser or user.is_staff

@login_required
@user_passes_test(is_admin)
def admin_dashboard_view(request):
    rentals = Rental.objects.select_related('item').all()

    monthly_data = defaultdict(lambda: defaultdict(int))  # {月: {品目: 貸出台数（個数）}}
    all_months_set = set()
    all_items_set = set()

    for rental in rentals:
        month = rental.rental_date.strftime('%Y-%m')
        item_name = rental.item.name
        monthly_data[month][item_name] += rental.quantity  # ← 件数でなく数量を加算
        all_months_set.add(month)
        all_items_set.add(item_name)

    labels = sorted(all_months_set)
    item_names = sorted(all_items_set)

    color_palette = [
        'rgba(255, 99, 132, 0.7)',
        'rgba(54, 162, 235, 0.7)',
        'rgba(255, 206, 86, 0.7)',
        'rgba(75, 192, 192, 0.7)',
        'rgba(153, 102, 255, 0.7)',
        'rgba(255, 159, 64, 0.7)',
    ]

    datasets = []
    for idx, item_name in enumerate(item_names):
        data = [monthly_data[month].get(item_name, 0) for month in labels]
        datasets.append({
            'label': item_name,
            'data': data,
            'backgroundColor': color_palette[idx % len(color_palette)],
            'borderColor': color_palette[idx % len(color_palette)].replace('0.7', '1'),
            'borderWidth': 1,
        })

    monthly_total_counts = [sum(monthly_data[month].values()) for month in labels]

    return render(request, 'inventory/admin_dashboard.html', {
        'labels': labels,
        'datasets': datasets,
        'data': monthly_total_counts,
    })

@login_required
def user_dashboard_view(request):
    items = InventoryItem.objects.filter(added_by=request.user)
    rentals = Rental.objects.filter(user=request.user).select_related('item')

    rental_active_counts = defaultdict(lambda: defaultdict(int))
    current_rentals_count = defaultdict(int)  # 品目別：現在貸出中数
    all_months_set = set()
    all_items_set = set()

    today = datetime.today()
    this_month = datetime(today.year, today.month, 1)

    for rental in rentals:
        item_name = rental.item.name
        all_items_set.add(item_name)

        if not rental.rental_date:
            continue

        start_date = datetime(rental.rental_date.year, rental.rental_date.month, 1)

        if rental.return_date:
            return_month = datetime(rental.return_date.year, rental.return_date.month, 1)
            end_date = return_month
        else:
            end_date = this_month
            # 現在貸出中の品目を集計
            current_rentals_count[item_name] += rental.quantity

        current = start_date
        while current <= end_date:
            month_str = current.strftime('%Y-%m')
            rental_active_counts[month_str][item_name] += rental.quantity
            all_months_set.add(month_str)

            if current.month == 12:
                current = datetime(current.year + 1, 1, 1)
            else:
                current = datetime(current.year, current.month + 1, 1)

    # 月を連続した時系列で整列
    sorted_months_dt = sorted([datetime.strptime(m, '%Y-%m') for m in all_months_set])
    if not sorted_months_dt:
        return render(request, 'inventory/user_dashboard.html', {
            'labels': [], 'datasets': [], 'data': [],
            'current_rentals': {}, 'rentals': rentals, 'items': items,
        })

    start = sorted_months_dt[0]
    end = sorted_months_dt[-1]
    current = datetime(start.year, start.month, 1)
    all_months = []

    while current <= end:
        all_months.append(current.strftime('%Y-%m'))
        if current.month == 12:
            current = datetime(current.year + 1, 1, 1)
        else:
            current = datetime(current.year, current.month + 1, 1)

    # 月別グラフ用データ構築
    item_names = sorted(all_items_set)
    datasets = []
    color_palette = [
        'rgba(255, 99, 132, 0.7)', 'rgba(54, 162, 235, 0.7)',
        'rgba(255, 206, 86, 0.7)', 'rgba(75, 192, 192, 0.7)',
        'rgba(153, 102, 255, 0.7)', 'rgba(255, 159, 64, 0.7)',
        'rgba(199, 199, 199, 0.7)', 'rgba(83, 102, 255, 0.7)',
        'rgba(255, 102, 255, 0.7)', 'rgba(102, 255, 204, 0.7)',
    ]

    for idx, item in enumerate(item_names):
        data = [rental_active_counts[month][item] for month in all_months]
        datasets.append({
            'label': item,
            'data': data,
            'backgroundColor': color_palette[idx % len(color_palette)],
            'borderColor': color_palette[idx % len(color_palette)].replace('0.7', '1'),
            'borderWidth': 1,
        })

    # 月ごとの合計貸出中数（全品目）
    monthly_total_counts = [sum(rental_active_counts[month].values()) for month in all_months]

    return render(request, 'inventory/user_dashboard.html', {
        'labels': all_months,
        'datasets': datasets,
        'data': monthly_total_counts,
        'current_rentals': dict(current_rentals_count),  # 現在貸出中の品目と数量
        'rentals': rentals,
        'items': items,
    })

@login_required
def redirect_after_login(request):
    if request.user.is_staff:
        return redirect('admin_dashboard')  
    else:
        return redirect('user_dashboard')

# 一覧表示ビュー
@login_required
def item_list(request):
    if request.user.is_staff:
        return HttpResponseForbidden("管理者ユーザーはこのページにアクセスできません。")

    form = ItemSearchForm(request.GET or None)  # GETパラメータでフォームを初期化
    items = InventoryItem.objects.filter(is_available=True)  # 利用可能な備品だけ取得

    if form.is_valid():
        query = form.cleaned_data.get('query')
        stock_filter = form.cleaned_data.get('stock_filter')

        if query:
            items = items.filter(
                Q(name__icontains=query) | Q(description__icontains=query)
            )

        if stock_filter == 'in_stock':
            items = items.filter(quantity__gt=0)
        elif stock_filter == 'out_of_stock':
            items = items.filter(quantity__lte=0)

    context = {
        'form': form,
        'items': items,
    }
    return render(request, 'inventory/item_list.html', context)

# 詳細表示ビュー
@login_required
def item_detail(request, pk):
    if request.user.is_staff:
        return HttpResponseForbidden("管理者ユーザーはこのページにアクセスできません。")
    
    item = get_object_or_404(InventoryItem, pk=pk)
    return render(request, 'inventory/item_detail.html', {'item': item})

# 登録
@login_required
@user_passes_test(is_general_user)
def item_create(request):
    if request.method == 'POST':
        form = InventoryItemForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.added_by = request.user
            item.save()
            return redirect('item_detail', pk=item.pk)  # 備品一覧にリダイレクト
    else:
        form = InventoryItemForm()
    return render(request, 'inventory/item_form.html', {'form': form, 'title': '備品登録'})

# 編集
@login_required
def item_update(request, pk):
    item = get_object_or_404(InventoryItem, pk=pk, added_by=request.user)
    if request.method == 'POST':
        form = InventoryItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return redirect('user_dashboard')
    else:
        form = InventoryItemForm(instance=item)
    return render(request, 'inventory/item_form.html', {'form': form, 'title': '備品編集'})

# 削除（確認画面あり）
@login_required
def item_delete(request, pk):
    item = get_object_or_404(InventoryItem, pk=pk, added_by=request.user)
    if request.method == 'POST':
        item.delete()
        return redirect('user_dashboard')
    return render(request, 'inventory/item_confirm_delete.html', {'item': item})

@login_required
def edit_item(request, pk):
    item = get_object_or_404(InventoryItem, pk=pk)

    # 一般ユーザーは自分が登録したものだけ編集可能
    if not request.user.is_superuser and item.added_by!= request.user:
        return redirect('user_dashboard')  # 不正アクセス防止

    if request.method == 'POST':
        form = InventoryItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return redirect('user_dashboard') 
    else:
        form = InventoryItemForm(instance=item)

    return render(request, 'inventory/edit_item.html', {'form': form, 'item': item})

class InventoryItemDeleteView(DeleteView):
    model = InventoryItem
    template_name = 'inventory/item_confirm_delete.html'
    success_url = reverse_lazy('items_list')

    def dispatch(self, request, *args, **kwargs):
        obj = self.get_object()
        if obj.added_by != request.user:
            raise PermissionDenied
        return super().dispatch(request, *args, **kwargs)
    
@login_required
def rental_create(request, item_id=None):
    item = None
    if item_id:
        item = get_object_or_404(InventoryItem, pk=item_id)

    if request.method == 'POST':
        # item を initial に渡す（cleanメソッドで使用）
        form = RentalForm(request.POST, initial={'item': item})
        if form.is_valid():
            rental = form.save(commit=False)
            rental.item = item
            rental.user = request.user
            rental.status = 'borrowed'

            if rental.quantity > item.quantity:
                messages.error(request, f"在庫数（{item.quantity}個）より多くは貸し出せません。")
                return redirect('user_dashboard')

            rental.save()
            item.quantity -= rental.quantity
            item.save()

            messages.success(
                request,
                f"{rental.item.name} を貸し出しました（予定返却日: {rental.expected_return_date}）。"
            )
            return redirect('user_dashboard')
    else:
        # GETリクエスト時にも item を initial に渡す
        form = RentalForm(initial={'item': item})

    return render(request, 'inventory/rental_create.html', {'form': form, 'item': item})

@login_required
def rental_list(request):
    rentals = Rental.objects.all().order_by('-rental_date')
    return render(request, 'inventory/rental_list.html', {'rentals': rentals})

@login_required
def return_item(request, rental_id):
    rental = get_object_or_404(Rental, id=rental_id, user=request.user)

    if request.method == 'POST' and rental.status == 'borrowed':
        if rental.quantity > 1:
            rental.quantity -= 1
            rental.save()
            
            ReturnLog.objects.create(rental=rental, returned_quantity=1, returned_by=request.user)
            messages.success(request, f"{rental.item.name}を1個返却しました。残り{rental.quantity}個貸出中です。")
        else:
            rental.quantity = 0
            rental.status = 'returned'
            rental.return_date = timezone.now().date()
            rental.save()
            
            ReturnLog.objects.create(rental=rental, returned_quantity=1, returned_by=request.user)
            messages.success(request, f"{rental.item.name}をすべて返却しました。")

    return redirect('user_dashboard')

@login_required
def all_rental_history_view(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("このページは管理者のみアクセス可能です。")
    
    rentals = Rental.objects.select_related('user', 'item').order_by('-rental_date')

    # フィルター用パラメータを取得
    user_id = request.GET.get('user')
    item_id = request.GET.get('item')
    status = request.GET.get('status')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if user_id:
        try:
            user_id_int = int(user_id)
            rentals = rentals.filter(user_id=user_id_int)
        except ValueError:
            pass

    if item_id:
        try:
            item_id_int = int(item_id)
            rentals = rentals.filter(item_id=item_id_int)
        except ValueError:
            pass

    if status:
        rentals = rentals.filter(status=status)

    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            rentals = rentals.filter(rental_date__gte=start_date_obj)
        except ValueError:
            pass

    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
            rentals = rentals.filter(rental_date__lt=end_date_obj)  # "翌日の0:00"未満を対象に含める
        except ValueError:
            pass

    users = User.objects.all()
    items = InventoryItem.objects.all()

    return render(request, 'inventory/all_rental_history.html', {
        'rentals': rentals,
        'users': users,
        'items': items,
        'selected_user': user_id or '',
        'selected_item': item_id or '',
        'selected_status': status or '',
        'start_date': start_date or '',
        'end_date': end_date or '',
    })

@login_required
def export_rentals_csv(request):
    rentals = Rental.objects.filter(user=request.user)

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="rental_history.csv"'

    response.write('\ufeff')  # Excel用のBOM

    writer = csv.writer(response)
    writer.writerow(['アイテム名', '数量', '貸出日', '返却予定日', '返却日', 'ステータス'])

    for rental in rentals:
        writer.writerow([
            rental.item.name,
            rental.quantity,
            rental.rental_date.strftime('%Y-%m-%d') if rental.rental_date else '',
            rental.expected_return_date.strftime('%Y-%m-%d') if rental.expected_return_date else '',
            rental.return_date.strftime('%Y-%m-%d') if rental.return_date else '未返却',
            dict(Rental.STATUS_CHOICES).get(rental.status, rental.status),
        ])

    return response

@login_required
def export_rentals_excel(request):
    rentals = Rental.objects.filter(user=request.user)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rental History"

    headers = ['アイテム名', '数量', '貸出日', '返却予定日', '返却日', 'ステータス']
    ws.append(headers)

    for rental in rentals:
        ws.append([
            rental.item.name,
            rental.quantity,
            rental.rental_date.strftime("%Y/%m/%d") if rental.rental_date else '',
            rental.expected_return_date.strftime("%Y/%m/%d") if rental.expected_return_date else '',
            rental.return_date.strftime("%Y/%m/%d") if rental.return_date else '未返却',
            dict(Rental.STATUS_CHOICES).get(rental.status, rental.status),
        ])

    for i in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="rental_history.xlsx"'
    wb.save(response)
    return response

@login_required
def export_rentals_pdf(request):
    rentals = Rental.objects.filter(user=request.user).order_by('-rental_date')

    base_url = request.build_absolute_uri('/')
    context = {
        'rentals': rentals,
        'user': request.user,
        'base_url': base_url,
    }

    # HTML をテンプレートから生成
    html_string = render_to_string('inventory/rental_history_pdf.html', context)

    # WeasyPrint のログ出力を有効化（Render のログタブで確認可能）
    logger = logging.getLogger("weasyprint")
    logger.setLevel(logging.DEBUG)
    logger.addHandler(logging.StreamHandler())

    # フォントファイルのパス
    font_path = os.path.join(settings.STATIC_ROOT, 'fonts', 'NotoSansCJKjp-Regular.otf')

    # カスタムCSS（フォント埋め込み）
    css = CSS(string=f"""
        @font-face {{
            font-family: 'Noto Sans CJK JP';
            src: url('file://{font_path}') format('opentype');
        }}
        body {{
            font-family: 'Noto Sans CJK JP', sans-serif;
            font-size: 12px;
        }}
        h1 {{
            text-align: center;
            margin-bottom: 20px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
        }}
        th, td {{
            border: 1px solid #000;
            padding: 8px;
            text-align: left;
        }}
        th {{
            background-color: #f2f2f2;
        }}
    """)

    # PDF を生成
    html = HTML(string=html_string, base_url=base_url)
    pdf = html.write_pdf(stylesheets=[css])

    # PDF をレスポンスとして返す
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="rental_history.pdf"'
    return response

@staff_member_required
def export_all_rentals_csv(request):
    rentals = Rental.objects.select_related('user', 'item').all()

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="all_rental_history.csv"'
    response.write('\ufeff')  

    writer = csv.writer(response)
    writer.writerow(['ユーザー名', 'アイテム名', '数量', '貸出日', '返却予定日', '返却日', 'ステータス'])

    for rental in rentals:
        writer.writerow([
            rental.user.username,
            rental.item.name,
            rental.quantity,
            rental.rental_date.strftime("%Y/%m/%d") if rental.rental_date else '',
            rental.expected_return_date.strftime("%Y/%m/%d") if rental.expected_return_date else '',
            rental.return_date.strftime("%Y/%m/%d") if rental.return_date else '未返却',
            dict(Rental.STATUS_CHOICES).get(rental.status, rental.status),
        ])

    return response


@staff_member_required
def export_all_rentals_excel(request):
    rentals = Rental.objects.select_related('user', 'item').all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Rental History"

    headers = ['ユーザー名', 'アイテム名', '数量', '貸出日', '返却予定日', '返却日', 'ステータス']
    ws.append(headers)

    for rental in rentals:
        ws.append([
            rental.user.username,
            rental.item.name,
            rental.quantity,
            rental.rental_date.strftime("%Y/%m/%d") if rental.rental_date else '',
            rental.expected_return_date.strftime("%Y/%m/%d") if rental.expected_return_date else '',
            rental.return_date.strftime("%Y/%m/%d") if rental.return_date else '未返却',
            dict(Rental.STATUS_CHOICES).get(rental.status, rental.status),
        ])

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="all_rental_history.xlsx"'
    wb.save(response)
    return response


@staff_member_required
def export_all_rentals_pdf(request):
    rentals = Rental.objects.all().order_by('user__username', '-rental_date')

    # ベースURLの取得
    base_url = request.build_absolute_uri('/')
    
    context = {
        'rentals': rentals,
        'current_user': request.user,
    }

    # HTMLをテンプレートから生成
    html_string = render_to_string('inventory/all_rentals_pdf.html', context)
    
    # WeasyPrintのログ設定（Renderのログタブで確認可能）
    logger = logging.getLogger("weasyprint")
    logger.setLevel(logging.DEBUG)
    logger.addHandler(logging.StreamHandler())
    
    # フォントファイルの絶対パスを取得
    font_path = os.path.join(settings.STATIC_ROOT, 'fonts', 'NotoSansCJKjp-Regular.otf')
    
    # カスタムCSS（日本語フォントの埋め込み）
    css = CSS(string=f"""
        @font-face {{
            font-family: 'Noto Sans CJK JP';
            src: url('file://{font_path}') format('opentype');
        }}
        body {{
            font-family: 'Noto Sans CJK JP', sans-serif;
            font-size: 12px;
        }}
        h2 {{
            text-align: center;
            margin-bottom: 30px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}
        th, td {{
            border: 1px solid #000;
            padding: 8px;
            text-align: center;
        }}
        th {{
            background-color: #f2f2f2;
        }}
    """)

    # PDFを生成
    html = HTML(string=html_string, base_url=base_url)
    pdf = html.write_pdf(stylesheets=[css])

    # PDFをレスポンスとして返す
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="all_rental_history.pdf"'
    return response

def get_monthly_rental_data(user):
    if user.is_staff:
        rentals = Rental.objects.all()
        # 管理者は貸出台数を「備品の数量（Sum）」で表示
        monthly_totals = rentals.annotate(month=TruncMonth('rental_date')) \
                                .values('month') \
                                .annotate(total_quantity=Sum('quantity')) \
                                .order_by('month')

        item_month_data = rentals.annotate(month=TruncMonth('rental_date')) \
                                 .values('item__name', 'month') \
                                 .annotate(total_quantity=Sum('quantity')) \
                                 .order_by('month')
    else:
        rentals = Rental.objects.filter(user=user)
        # 一般ユーザーは貸出台数を「借りた回数（Count）」で表示
        monthly_totals = rentals.annotate(month=TruncMonth('rental_date')) \
                                .values('month') \
                                .annotate(total_quantity=Count('id')) \
                                .order_by('month')

        item_month_data = rentals.annotate(month=TruncMonth('rental_date')) \
                                 .values('item__name', 'month') \
                                 .annotate(total_quantity=Count('id')) \
                                 .order_by('month')

    # 共通処理：月ごとのラベルとデータ作成
    labels = []
    data = []
    for entry in monthly_totals:
        month_label = entry['month'].strftime('%Y-%m')
        labels.append(month_label)
        data.append(entry['total_quantity'])

    # 品目別・月別データ整形
    item_data = defaultdict(lambda: [0] * len(labels))
    month_index = {label: i for i, label in enumerate(labels)}

    for entry in item_month_data:
        item_name = entry['item__name']
        month_label = entry['month'].strftime('%Y-%m')
        idx = month_index.get(month_label)
        if idx is not None:
            item_data[item_name][idx] = entry['total_quantity']

    # Chart.js datasets形式へ変換
    datasets = [{
        'label': item,
        'data': quantities,
        'backgroundColor': f'rgba({(i * 50) % 255}, {(i * 80) % 255}, {(i * 110) % 255}, 0.6)'
    } for i, (item, quantities) in enumerate(item_data.items())]

    return labels, data, datasets

def home(request):
    if request.user.is_authenticated:
        if request.user.is_staff:
            return redirect('admin_dashboard')  
        else:
            return redirect('login')  
    return redirect('login') 

class SignupView(View):
    def dispatch(self, request, *args, **kwargs):
        if not settings.DEBUG:
            return HttpResponseForbidden("本番環境では新規登録できません。")
        return super().dispatch(request, *args, **kwargs)
    